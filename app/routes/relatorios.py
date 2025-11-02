from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import io
import tempfile
import os

from app.core.database import get_db
from app.core.auth import get_current_user_from_cookie
from app.models.usuario import Usuario
from app.services.relatorio_service import RelatorioService

router = APIRouter(prefix="/api/relatorios", tags=["relatorios"])


@router.get("/mensal")
async def gerar_relatorio_mensal(
    ano: int = Query(..., description="Ano de referência"),
    mes: int = Query(..., ge=1, le=12, description="Mês de referência (1-12)"),
    proprietario_id: Optional[int] = Query(None, description="ID do proprietário (opcional)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_from_cookie)
):
    """
    Gera relatório mensal consolidado
    
    - **ano**: Ano de referência
    - **mes**: Mês de referência (1-12)
    - **proprietario_id**: Opcional - ID do proprietário para filtrar
    """
    try:
        relatorio = RelatorioService.gerar_relatorio_mensal(
            db=db,
            ano=ano,
            mes=mes,
            proprietario_id=proprietario_id
        )
        return relatorio
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")


@router.get("/proprietario/{proprietario_id}")
async def gerar_relatorio_proprietario(
    proprietario_id: int,
    ano: int = Query(..., description="Ano de referência"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Mês de referência (opcional)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_from_cookie)
):
    """
    Gera relatório específico de um proprietário
    
    - **proprietario_id**: ID do proprietário
    - **ano**: Ano de referência
    - **mes**: Opcional - Mês específico (se omitido, gera relatório anual)
    """
    try:
        relatorio = RelatorioService.gerar_relatorio_proprietario(
            db=db,
            proprietario_id=proprietario_id,
            ano=ano,
            mes=mes
        )
        
        if "erro" in relatorio:
            raise HTTPException(status_code=404, detail=relatorio["erro"])
        
        return relatorio
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")


@router.get("/anual")
async def gerar_relatorio_anual(
    ano: int = Query(..., description="Ano de referência"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_from_cookie)
):
    """
    Gera relatório anual consolidado
    
    - **ano**: Ano de referência
    """
    try:
        relatorio = RelatorioService.gerar_relatorio_anual(
            db=db,
            ano=ano
        )
        return relatorio
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")


@router.get("/comparativo")
async def gerar_relatorio_comparativo(
    ano1: int = Query(..., description="Primeiro ano"),
    ano2: int = Query(..., description="Segundo ano"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_from_cookie)
):
    """
    Gera relatório comparativo entre dois anos
    
    - **ano1**: Primeiro ano para comparação
    - **ano2**: Segundo ano para comparação
    """
    try:
        relatorio = RelatorioService.gerar_relatorio_comparativo(
            db=db,
            ano1=ano1,
            ano2=ano2
        )
        return relatorio
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")


@router.get("/dashboard")
async def obter_dados_dashboard(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_from_cookie)
):
    """
    Retorna dados agregados para o dashboard principal
    
    Inclui:
    - Estatísticas do mês atual
    - Estatísticas do ano atual
    - Comparação com mês anterior
    - Top 5 imóveis por receita
    """
    try:
        hoje = datetime.now()
        ano_atual = hoje.year
        mes_atual = hoje.month
        
        # Relatório do mês atual
        relatorio_mes_atual = RelatorioService.gerar_relatorio_mensal(
            db=db,
            ano=ano_atual,
            mes=mes_atual
        )
        
        # Relatório do mês anterior
        mes_anterior = mes_atual - 1 if mes_atual > 1 else 12
        ano_anterior = ano_atual if mes_atual > 1 else ano_atual - 1
        
        relatorio_mes_anterior = RelatorioService.gerar_relatorio_mensal(
            db=db,
            ano=ano_anterior,
            mes=mes_anterior
        )
        
        # Relatório anual
        relatorio_anual = RelatorioService.gerar_relatorio_anual(
            db=db,
            ano=ano_atual
        )
        
        # Calcular variação mensal
        from decimal import Decimal
        receita_mes_atual = Decimal(str(relatorio_mes_atual["resumo"]["total_recebido"]))
        receita_mes_anterior = Decimal(str(relatorio_mes_anterior["resumo"]["total_recebido"]))
        
        variacao_mensal = receita_mes_atual - receita_mes_anterior
        variacao_percentual = ((receita_mes_atual / receita_mes_anterior - 1) * 100) if receita_mes_anterior > 0 else Decimal('0')
        
        # Top 5 imóveis por receita do mês
        detalhamento = relatorio_mes_atual["detalhamento"]
        top_imoveis = sorted(
            detalhamento,
            key=lambda x: x["valores"]["total"],
            reverse=True
        )[:5]
        
        return {
            "mes_atual": {
                "periodo": relatorio_mes_atual["periodo"],
                "resumo": relatorio_mes_atual["resumo"]
            },
            "comparacao_mensal": {
                "variacao_absoluta": float(variacao_mensal),
                "variacao_percentual": float(variacao_percentual),
                "mes_anterior": {
                    "ano": ano_anterior,
                    "mes": mes_anterior,
                    "total_recebido": float(receita_mes_anterior)
                }
            },
            "anual": {
                "ano": ano_atual,
                "resumo": relatorio_anual["resumo"]
            },
            "top_imoveis": top_imoveis
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao obter dados do dashboard: {str(e)}")


@router.get("/exportar/pdf/mensal")
async def exportar_relatorio_mensal_pdf(
    ano: int = Query(..., description="Ano de referência"),
    mes: int = Query(..., ge=1, le=12, description="Mês de referência"),
    proprietario_id: Optional[int] = Query(None, description="ID do proprietário"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_from_cookie)
):
    """
    Exporta relatório mensal em formato PDF
    
    - **ano**: Ano de referência
    - **mes**: Mês de referência (1-12)
    - **proprietario_id**: Opcional - ID do proprietário para filtrar
    """
    try:
        # Importar aqui para evitar erro se reportlab não estiver instalado
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        
        # Gerar relatório
        relatorio = RelatorioService.gerar_relatorio_mensal(
            db=db,
            ano=ano,
            mes=mes,
            proprietario_id=proprietario_id
        )
        
        # Criar PDF em memória
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo customizado
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#135bec'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        periodo = relatorio["periodo"]
        titulo = f"Relatório Mensal - {periodo['mes_nome']}/{periodo['ano']}"
        elements.append(Paragraph(titulo, title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Resumo
        resumo = relatorio["resumo"]
        data_resumo = [
            ['Métrica', 'Valor'],
            ['Total de Aluguéis', str(resumo['total_alugueis'])],
            ['Aluguéis Pagos', str(resumo['alugueis_pagos'])],
            ['Aluguéis Pendentes', str(resumo['alugueis_pendentes'])],
            ['Total Esperado', f"R$ {resumo['total_esperado']:,.2f}"],
            ['Total Recebido', f"R$ {resumo['total_recebido']:,.2f}"],
            ['Total Pendente', f"R$ {resumo['total_pendente']:,.2f}"],
            ['Taxa de Recebimento', f"{resumo['taxa_recebimento']:.1f}%"]
        ]
        
        table_resumo = Table(data_resumo, colWidths=[8*cm, 8*cm])
        table_resumo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#135bec')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table_resumo)
        elements.append(Spacer(1, 1*cm))
        
        # Detalhamento por imóvel
        if relatorio["detalhamento"]:
            elements.append(Paragraph("Detalhamento por Imóvel", styles['Heading2']))
            elements.append(Spacer(1, 0.5*cm))
            
            data_detalhes = [['Imóvel', 'Status', 'Valor Total']]
            
            for item in relatorio["detalhamento"]:
                data_detalhes.append([
                    item['imovel_endereco'][:40],
                    item['status_pagamento'].upper(),
                    f"R$ {item['valores']['total']:,.2f}"
                ])
            
            table_detalhes = Table(data_detalhes, colWidths=[10*cm, 3*cm, 3*cm])
            table_detalhes.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#135bec')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table_detalhes)
        
        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"relatorio_mensal_{ano}_{mes:02d}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca reportlab não instalada. Execute: pip install reportlab"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar PDF: {str(e)}")


@router.get("/exportar/excel/mensal")
async def exportar_relatorio_mensal_excel(
    ano: int = Query(..., description="Ano de referência"),
    mes: int = Query(..., ge=1, le=12, description="Mês de referência"),
    proprietario_id: Optional[int] = Query(None, description="ID do proprietário"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user_from_cookie)
):
    """
    Exporta relatório mensal em formato Excel
    
    - **ano**: Ano de referência
    - **mes**: Mês de referência (1-12)
    - **proprietario_id**: Opcional - ID do proprietário para filtrar
    """
    try:
        # Importar aqui para evitar erro se openpyxl não estiver instalado
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Gerar relatório
        relatorio = RelatorioService.gerar_relatorio_mensal(
            db=db,
            ano=ano,
            mes=mes,
            proprietario_id=proprietario_id
        )
        
        # Criar workbook
        wb = Workbook()
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        # Estilos
        header_fill = PatternFill(start_color="135bec", end_color="135bec", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14, color="135bec")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        periodo = relatorio["periodo"]
        ws_resumo['A1'] = f"Relatório Mensal - {periodo['mes_nome']}/{periodo['ano']}"
        ws_resumo['A1'].font = title_font
        ws_resumo.merge_cells('A1:B1')
        
        # Resumo
        ws_resumo['A3'] = "Métrica"
        ws_resumo['B3'] = "Valor"
        ws_resumo['A3'].fill = header_fill
        ws_resumo['B3'].fill = header_fill
        ws_resumo['A3'].font = header_font
        ws_resumo['B3'].font = header_font
        
        resumo = relatorio["resumo"]
        row = 4
        for label, value in [
            ('Total de Aluguéis', resumo['total_alugueis']),
            ('Aluguéis Pagos', resumo['alugueis_pagos']),
            ('Aluguéis Pendentes', resumo['alugueis_pendentes']),
            ('Total Esperado', f"R$ {resumo['total_esperado']:,.2f}"),
            ('Total Recebido', f"R$ {resumo['total_recebido']:,.2f}"),
            ('Total Pendente', f"R$ {resumo['total_pendente']:,.2f}"),
            ('Taxa de Recebimento', f"{resumo['taxa_recebimento']:.1f}%")
        ]:
            ws_resumo[f'A{row}'] = label
            ws_resumo[f'B{row}'] = value
            ws_resumo[f'A{row}'].border = border
            ws_resumo[f'B{row}'].border = border
            row += 1
        
        # Ajustar largura das colunas
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 20
        
        # Criar aba de detalhamento
        if relatorio["detalhamento"]:
            ws_detalhes = wb.create_sheet("Detalhamento")
            
            # Cabeçalhos
            headers = ['Imóvel', 'Status', 'Vencimento', 'Aluguel', 'Condomínio', 
                      'IPTU', 'Luz', 'Água', 'Gás', 'Internet', 'Outros', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws_detalhes.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            # Dados
            for row, item in enumerate(relatorio["detalhamento"], 2):
                valores = item['valores']
                ws_detalhes.cell(row, 1, item['imovel_endereco'])
                ws_detalhes.cell(row, 2, item['status_pagamento'].upper())
                ws_detalhes.cell(row, 3, item['data_vencimento'])
                ws_detalhes.cell(row, 4, valores['aluguel'])
                ws_detalhes.cell(row, 5, valores['condominio'])
                ws_detalhes.cell(row, 6, valores['iptu'])
                ws_detalhes.cell(row, 7, valores['luz'])
                ws_detalhes.cell(row, 8, valores['agua'])
                ws_detalhes.cell(row, 9, valores['gas'])
                ws_detalhes.cell(row, 10, valores['internet'])
                ws_detalhes.cell(row, 11, valores['outros'])
                ws_detalhes.cell(row, 12, valores['total'])
                
                # Aplicar bordas
                for col in range(1, 13):
                    ws_detalhes.cell(row, col).border = border
            
            # Ajustar larguras
            ws_detalhes.column_dimensions['A'].width = 40
            for col in ['B', 'C']:
                ws_detalhes.column_dimensions[col].width = 15
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws_detalhes.column_dimensions[col].width = 12
        
        # Salvar em memória
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"relatorio_mensal_{ano}_{mes:02d}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Biblioteca openpyxl não instalada. Execute: pip install openpyxl"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar Excel: {str(e)}")
